#!/usr/bin/env pyhon
from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)
import openpyxl
import time

book = Workbook()
sheet = book.active

sheet['A1'] = 56
sheet['A2'] = 43

now = time.strftime("%x")
sheet['A3'] = now

book.save('book.xlsx')

rows = (
    (1, 2),
    (3, 4),
    (5, 6),
    (7, 8),
    (9, 10),
    (11, 12),
    (13, 14),
)

book2 = Workbook()
sheet2 = book2.active

for  row in rows:
    sheet2.append(row)

cell = sheet2.cell(row=8, column=2)
cell.value = '=SUM(A1:B6)'
cell.font = cell.font.copy(bold=True)

sheet2.sheet_properties.tabColor = "000275"

book2.save('numbers3.xlsx')


book3 = openpyxl.load_workbook('numbers3.xlsx')

sheet3 = book3.active

a1 = sheet3['A1']
a2 = sheet3['A2']
a3 = sheet3.cell(row=3, column=1)

print(a1.value)
print(a2.value) 
print(a3.value)

data = Reference(sheet3, min_col=2, min_row=1, max_col=2, max_row=7)
categs = Reference(sheet3, min_col=1, min_row=1, max_row=7)

chart = BarChart()
chart.add_data(data=data)
chart.set_categories(categs)

chart.legend = None
chart.y_axis.majorGridlines = None
chart.varyColors = True
chart.title = "Numbers"

sheet3.add_chart(chart, "A10")    

book3.save("numbers3.xlsx")